from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, date
from functools import wraps
from src.recognition.live_recognizer import FaceRecognizer
import os
import io
import csv
import threading
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

#------------------------------------------------------------------------------Config
BASE_DIR = Path(__file__).parent
CAM_FRAME_PATH = BASE_DIR / 'cam_frames'
EMBEDDING_MODEL_PATH = BASE_DIR / 'models' / 'w600k_r50.onnx'
CLASSIFIER_MODEL_PATH = BASE_DIR / 'models' / 'knn_model.pkl'
RECOGNITION_GPU_CTX = 0 
RECOGNITION_MIN_CONFIDENCE = 0.5
RECOGNITION_MIN_AGREEMENT = 3 
EXPECTED_FRAMES = 5

ALLOWED_EXTENSIONS = {
    "jpg",
    "jpeg",
    "png"
}

FIRST_FRAME_SUFFIX = "_01.jpg"

app = Flask(__name__)
app.config['SECRET_KEY'] = 'change-this-secret-key-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(os.path.dirname(__file__), 'attendance.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

CAM_FRAME_PATH.mkdir(
    parents=True,
    exist_ok=True
)

recognizer = FaceRecognizer(

    cam_frame_path=CAM_FRAME_PATH,

    embedding_model_path=EMBEDDING_MODEL_PATH,

    classifier_model_path=CLASSIFIER_MODEL_PATH,

    gpu_ctx_id=RECOGNITION_GPU_CTX
)

YEAR_CHOICES = ['1st Year', '2nd Year', '3rd Year', '4th Year']

# -----------------------------------------------------------------------------------DataBase Model

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='teacher')  # admin / teacher

    def set_password(self, pw):
        self.password_hash = generate_password_hash(pw)

    def check_password(self, pw):
        return check_password_hash(self.password_hash, pw)


class Department(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    subjects = db.relationship('Subject', backref='department', lazy=True)
    students = db.relationship('Student', backref='department', lazy=True)


class Subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    code = db.Column(db.String(20), nullable=False, unique=True)
    department_id = db.Column(db.Integer, db.ForeignKey('department.id'), nullable=False)
    semester = db.Column(db.String(20))


class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    roll_no = db.Column(db.String(30), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120))
    phone = db.Column(db.String(20))
    department_id = db.Column(db.Integer, db.ForeignKey('department.id'), nullable=False)
    year = db.Column(db.String(20), nullable=False, default='1st Year')
    attendance_records = db.relationship('Attendance', backref='student', lazy=True, cascade="all, delete-orphan")


class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    date = db.Column(db.Date, nullable=False, default=date.today)
    status = db.Column(db.String(10), nullable=False)  # Present / Absent / Late
    __table_args__ = (db.UniqueConstraint('student_id', 'subject_id', 'date', name='unique_student_subject_date'),)
    subject = db.relationship('Subject')


#--------------------------------------------------------------------------------Authentication Functions
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in first.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Admin access required.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def allowed_file(filename):

    return (
        "." in filename
        and
        filename.rsplit(".", 1)[1].lower()
        in ALLOWED_EXTENSIONS
    )


def clear_frame_directory():
    for name in os.listdir(CAM_FRAME_PATH):
        path = CAM_FRAME_PATH / name
        if path.is_file():
            path.unlink()
    print(f"[upload] cleared old images from {CAM_FRAME_PATH}")

@app.context_processor
def inject_user():
    return dict(current_user_role=session.get('role'), current_username=session.get('username'),
                year_choices=YEAR_CHOICES)

#--------------------------------------------------------------------------------Authentication Routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            flash(f'Welcome back, {user.username}!', 'success')
            return redirect(url_for('dashboard'))
        flash('Invalid username or password.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    # Only allowed if no users exist yet (initial setup) or by an admin
    if User.query.count() > 0 and session.get('role') != 'admin':
        flash('Registration is restricted. Contact an admin.', 'danger')
        return redirect(url_for('login'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form.get('role', 'teacher')
        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'danger')
        else:
            u = User(username=username, role=role)
            u.set_password(password)
            db.session.add(u)
            db.session.commit()
            flash('Account created. Please log in.', 'success')
            return redirect(url_for('login'))
    return render_template('register.html')


# ---------------------------------------------------------------------------------Dashboard and Management Routes
@app.route('/')
@login_required
def dashboard():
    total_students = Student.query.count()
    total_subjects = Subject.query.count()
    total_departments = Department.query.count()
    today = date.today()
    today_present = Attendance.query.filter_by(date=today, status='Present').count()
    today_absent = Attendance.query.filter_by(date=today, status='Absent').count()
    today_late = Attendance.query.filter_by(date=today, status='Late').count()
    subjects = Subject.query.all()
    return render_template('dashboard.html',
                            total_students=total_students,
                            total_subjects=total_subjects,
                            total_departments=total_departments,
                            today_present=today_present,
                            today_absent=today_absent,
                            today_late=today_late,
                            subjects=subjects,
                            today=today)


#-------------------------------------------------------------------------------------Departments Routes
@app.route('/departments')
@login_required
def departments():
    depts = Department.query.all()
    return render_template('departments.html', departments=depts)

@app.route('/departments/add', methods=['POST'])
@login_required
@admin_required
def add_department():
    name = request.form['name'].strip()
    if name:
        if Department.query.filter_by(name=name).first():
            flash('Department already exists.', 'danger')
        else:
            db.session.add(Department(name=name))
            db.session.commit()
            flash('Department added.', 'success')
    return redirect(url_for('departments'))

@app.route('/departments/delete/<int:id>')
@login_required
@admin_required
def delete_department(id):
    dept = Department.query.get_or_404(id)
    if dept.subjects or dept.students:
        flash('Cannot delete department with existing subjects or students.', 'danger')
    else:
        db.session.delete(dept)
        db.session.commit()
        flash('Department deleted.', 'success')
    return redirect(url_for('departments'))


#--------------------------------------------------------------------------------Subjects Routes
@app.route('/subjects')
@login_required
def subjects():
    all_subjects = Subject.query.all()
    depts = Department.query.all()
    return render_template('subjects.html', subjects=all_subjects, departments=depts)

@app.route('/subjects/add', methods=['POST'])
@login_required
@admin_required
def add_subject():
    name = request.form['name'].strip()
    code = request.form['code'].strip()
    department_id = request.form['department_id']
    semester = request.form.get('semester', '').strip()
    if Subject.query.filter_by(code=code).first():
        flash('Subject code already exists.', 'danger')
    else:
        db.session.add(Subject(name=name, code=code, department_id=department_id, semester=semester))
        db.session.commit()
        flash('Subject added.', 'success')
    return redirect(url_for('subjects'))

@app.route('/subjects/delete/<int:id>')
@login_required
@admin_required
def delete_subject(id):
    subject = Subject.query.get_or_404(id)
    if Attendance.query.filter_by(subject_id=id).first():
        flash('Cannot delete subject with existing attendance records.', 'danger')
    else:
        db.session.delete(subject)
        db.session.commit()
        flash('Subject deleted.', 'success')
    return redirect(url_for('subjects'))

#--------------------------------------------------------------------------------Students Routes
@app.route('/students')
@login_required
def students():
    department_filter = request.args.get('department_id', type=int)
    year_filter = request.args.get('year', '')
    query = Student.query
    if department_filter:
        query = query.filter_by(department_id=department_filter)
    if year_filter:
        query = query.filter_by(year=year_filter)
    all_students = query.order_by(Student.name).all()
    all_departments = Department.query.all()
    return render_template('students.html', students=all_students, departments=all_departments,
                            selected_department=department_filter, selected_year=year_filter)

@app.route('/students/add', methods=['POST'])
@login_required
@admin_required
def add_student():
    roll_no = request.form['roll_no'].strip()
    name = request.form['name'].strip()
    email = request.form.get('email', '').strip()
    phone = request.form.get('phone', '').strip()
    department_id = request.form['department_id']
    year = request.form['year']
    if Student.query.filter_by(roll_no=roll_no).first():
        flash('Roll number already exists.', 'danger')
    else:
        db.session.add(Student(roll_no=roll_no, name=name, email=email, phone=phone,
                                department_id=department_id, year=year))
        db.session.commit()
        flash('Student added.', 'success')
    return redirect(url_for('students'))

@app.route('/students/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_student(id):
    student = Student.query.get_or_404(id)
    departments_list = Department.query.all()
    if request.method == 'POST':
        student.roll_no = request.form['roll_no'].strip()
        student.name = request.form['name'].strip()
        student.email = request.form.get('email', '').strip()
        student.phone = request.form.get('phone', '').strip()
        student.department_id = request.form['department_id']
        student.year = request.form['year']
        db.session.commit()
        flash('Student updated.', 'success')
        return redirect(url_for('students'))
    return render_template('edit_student.html', student=student, departments=departments_list)

@app.route('/students/delete/<int:id>')
@login_required
@admin_required
def delete_student(id):
    student = Student.query.get_or_404(id)
    db.session.delete(student)
    db.session.commit()
    flash('Student deleted.', 'success')
    return redirect(url_for('students'))

@app.route('/students/export')
@login_required
def export_students():
    department_filter = request.args.get('department_id', type=int)
    year_filter = request.args.get('year', '')
    query = Student.query
    if department_filter:
        query = query.filter_by(department_id=department_filter)
    if year_filter:
        query = query.filter_by(year=year_filter)
    all_students = query.order_by(Student.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Students'

    headers = ['Roll No.', 'Name', 'Department', 'Year', 'Email', 'Phone']
    dept_name = None
    if department_filter:
        d = Department.query.get(department_filter)
        dept_name = d.name if d else None

    title_bits = ['Student List']
    if dept_name:
        title_bits.append(dept_name)
    if year_filter:
        title_bits.append(year_filter)
    title = ' — '.join(title_bits) if len(title_bits) > 1 else 'Student List — All'
    subtitle = f"Generated {date.today().strftime('%d %B %Y')}"
    start_row = add_title(ws, title, len(headers), subtitle)

    for col, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col, value=h)
    style_header_row(ws, start_row, len(headers))

    row = start_row + 1
    for s in all_students:
        values = [s.roll_no, s.name, s.department.name, s.year, s.email or '', s.phone or '']
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = THIN_BORDER
        row += 1

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    autofit_columns(ws, [14, 24, 24, 12, 28, 16])

    fname_bits = [dept_name.replace(' ', '_') if dept_name else 'all', year_filter.replace(' ', '_') if year_filter else '']
    filename = f"students_{'_'.join([b for b in fname_bits if b])}_{date.today().isoformat()}.xlsx"
    return send_workbook(wb, filename)

@app.route('/students/export.csv')
@login_required
def export_students_csv():
    department_filter = request.args.get('department_id', type=int)
    year_filter = request.args.get('year', '')
    query = Student.query
    if department_filter:
        query = query.filter_by(department_id=department_filter)
    if year_filter:
        query = query.filter_by(year=year_filter)
    all_students = query.order_by(Student.name).all()

    dept_name = None
    if department_filter:
        d = Department.query.get(department_filter)
        dept_name = d.name if d else None

    rows = [['Roll No.', 'Name', 'Department', 'Year', 'Email', 'Phone']]
    for s in all_students:
        rows.append([s.roll_no, s.name, s.department.name, s.year, s.email or '', s.phone or ''])

    fname_bits = [dept_name.replace(' ', '_') if dept_name else 'all', year_filter.replace(' ', '_') if year_filter else '']
    filename = f"students_{'_'.join([b for b in fname_bits if b])}_{date.today().isoformat()}.csv"
    return send_csv(rows, filename)

# -----------------------------------------------------------------------------------Attendance Routes
@app.route('/attendance', methods=['GET', 'POST'])
@login_required
def attendance():

    # All departments and subjects for dropdowns
    all_departments = Department.query.order_by(Department.name).all()
    all_subjects = Subject.query.order_by(Subject.name).all()

    department_id = (
        request.args.get('department_id', type=int)
        or request.form.get('department_id', type=int)
    )

    subject_id = (
        request.args.get('subject_id', type=int)
        or request.form.get('subject_id', type=int)
    )

    year_filter = (
        request.args.get('year', '')
        or request.form.get('year', '')
    )

    selected_date_str = (
        request.args.get('date')
        or request.form.get('date')
        or date.today().isoformat()
    )

    selected_date = datetime.strptime(
        selected_date_str,
        '%Y-%m-%d'
    ).date()

    # Save Attendance

    if request.method == 'POST' and subject_id:

        subject = Subject.query.get_or_404(subject_id)

        # Make sure selected department matches subject department
        if department_id and subject.department_id != department_id:
            flash('Selected subject does not belong to selected department.', 'danger')
            return redirect(url_for('attendance'))

        student_query = Student.query.filter_by(
            department_id=subject.department_id
        )

        if year_filter:
            student_query = student_query.filter_by(
                year=year_filter
            )

        eligible_students = student_query.all()

        for student in eligible_students:

            status = request.form.get(
                f'status_{student.id}'
            )

            if not status:
                continue

            existing = Attendance.query.filter_by(
                student_id=student.id,
                subject_id=subject_id,
                date=selected_date
            ).first()

            if existing:

                # Update existing attendance
                existing.status = status

            else:

                # Create new attendance
                db.session.add(
                    Attendance(
                        student_id=student.id,
                        subject_id=subject_id,
                        date=selected_date,
                        status=status
                    )
                )

        db.session.commit()

        flash(
            'Attendance saved successfully.',
            'success'
        )

        return redirect(
            url_for(
                'attendance',
                department_id=subject.department_id,
                subject_id=subject_id,
                year=year_filter,
                date=selected_date_str
            )
        )

    students_list = []
    existing_records = {}

    if subject_id:

        subject = Subject.query.get_or_404(subject_id)

        # Validate department
        if department_id and subject.department_id != department_id:
            students_list = []

        else:

            student_query = Student.query.filter_by(
                department_id=subject.department_id
            )

            if year_filter:
                student_query = student_query.filter_by(
                    year=year_filter
                )

            students_list = student_query.order_by(
                Student.roll_no
            ).all()

            # Existing attendance for selected date + subject
            records = Attendance.query.filter_by(
                subject_id=subject_id,
                date=selected_date
            ).all()

            existing_records = {
                record.student_id: record.status
                for record in records
            }

    return render_template(
        'attendance.html',

        departments=all_departments,
        subjects=all_subjects,
        students=students_list,

        selected_department=department_id,
        selected_subject=subject_id,
        selected_year=year_filter,
        selected_date=selected_date_str,

        existing_records=existing_records
    )


# ---------------------------------------------------------------------Excel and CSV Export Utilities
HEADER_FILL = PatternFill(start_color='12233D', end_color='12233D', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='12233D')
SUB_FONT = Font(name='Calibri', italic=True, size=9, color='6B7280')
THIN_BORDER = Border(*(Side(style='thin', color='DDE3EC'),) * 4)
CENTER = Alignment(horizontal='center', vertical='center')

def style_header_row(ws, row_idx, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def autofit_columns(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_title(ws, title, num_cols, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal='left')
    row = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        sc = ws.cell(row=2, column=1, value=subtitle)
        sc.font = SUB_FONT
        row = 3
    return row + 1  # next free row (with a blank spacer row)

def send_workbook(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def send_csv(rows, filename):
    """rows: list of lists/tuples, first row is the header."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in rows:
        writer.writerow(row)
    mem = io.BytesIO(buf.getvalue().encode('utf-8-sig'))  # BOM so Excel opens UTF-8 cleanly
    mem.seek(0)
    return send_file(mem, as_attachment=True, download_name=filename, mimetype='text/csv')

#--------------------------------------------------------------------------------Reports Routes
def _attendance_summary(student_id, subject_id, start_date, end_date):
    q = Attendance.query.filter_by(student_id=student_id, subject_id=subject_id)
    if start_date:
        q = q.filter(Attendance.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        q = q.filter(Attendance.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    records = q.all()
    total = len(records)
    present = len([r for r in records if r.status == 'Present'])
    absent = len([r for r in records if r.status == 'Absent'])
    late = len([r for r in records if r.status == 'Late'])
    percentage = round((present + late) / total * 100, 1) if total > 0 else 0
    return total, present, absent, late, percentage


def _report_students(subject_id, year_filter):
    subject = Subject.query.get_or_404(subject_id)
    q = Student.query.filter_by(department_id=subject.department_id)
    if year_filter:
        q = q.filter_by(year=year_filter)
    return subject, q.order_by(Student.roll_no).all()

@app.route('/reports')
@login_required
def reports():
    all_subjects = Subject.query.all()
    subject_id = request.args.get('subject_id', type=int)
    year_filter = request.args.get('year', '')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    report_data = []
    if subject_id:
        _, students_list = _report_students(subject_id, year_filter)
        for student in students_list:
            total, present, absent, late, percentage = _attendance_summary(student.id, subject_id, start_date, end_date)
            report_data.append({
                'student': student, 'total': total, 'present': present,
                'absent': absent, 'late': late, 'percentage': percentage
            })

    return render_template('reports.html', subjects=all_subjects, report_data=report_data,
                            selected_subject=subject_id, selected_year=year_filter,
                            start_date=start_date, end_date=end_date)


@app.route('/reports/export')
@login_required
def export_report():
    subject_id = request.args.get('subject_id', type=int)
    year_filter = request.args.get('year', '')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not subject_id:
        flash('Select a subject first.', 'danger')
        return redirect(url_for('reports'))

    subject, students_list = _report_students(subject_id, year_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'

    headers = ['Roll No.', 'Name', 'Year', 'Total Days', 'Present', 'Absent', 'Late', 'Attendance %']
    date_range = ''
    if start_date or end_date:
        date_range = f" ({start_date or '…'} to {end_date or '…'})"
    year_bit = f" · {year_filter}" if year_filter else ''
    subtitle = f"{subject.code} - {subject.name}{year_bit}{date_range} · Generated {date.today().strftime('%d %B %Y')}"
    start_row = add_title(ws, 'Attendance Report', len(headers), subtitle)

    for col, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col, value=h)
    style_header_row(ws, start_row, len(headers))

    green_fill = PatternFill(start_color='E6F7EF', end_color='E6F7EF', fill_type='solid')
    red_fill = PatternFill(start_color='FDECEA', end_color='FDECEA', fill_type='solid')

    row = start_row + 1
    for student in students_list:
        total, present, absent, late, percentage = _attendance_summary(student.id, subject_id, start_date, end_date)

        values = [student.roll_no, student.name, student.year, total, present, absent, late, percentage / 100]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = THIN_BORDER
            cell.alignment = CENTER if col >= 4 else Alignment(horizontal='left')
        pct_cell = ws.cell(row=row, column=8)
        pct_cell.number_format = '0.0%'
        pct_cell.fill = green_fill if percentage >= 75 else red_fill
        pct_cell.font = Font(bold=True, color='1A9E6B' if percentage >= 75 else 'D9534F')
        row += 1

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    autofit_columns(ws, [12, 24, 12, 12, 10, 10, 8, 14])

    filename = f"attendance_report_{subject.code}_{date.today().isoformat()}.xlsx"
    return send_workbook(wb, filename)


@app.route('/reports/export.csv')
@login_required
def export_report_csv():
    subject_id = request.args.get('subject_id', type=int)
    year_filter = request.args.get('year', '')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not subject_id:
        flash('Select a subject first.', 'danger')
        return redirect(url_for('reports'))

    subject, students_list = _report_students(subject_id, year_filter)

    rows = [['Roll No.', 'Name', 'Year', 'Total Days', 'Present', 'Absent', 'Late', 'Attendance %']]
    for student in students_list:
        total, present, absent, late, percentage = _attendance_summary(student.id, subject_id, start_date, end_date)
        rows.append([student.roll_no, student.name, student.year, total, present, absent, late, f"{percentage}%"])

    filename = f"attendance_report_{subject.code}_{date.today().isoformat()}.csv"
    return send_csv(rows, filename)


@app.route('/reports/student/<int:id>')
@login_required
def student_report(id):
    student = Student.query.get_or_404(id)
    records = Attendance.query.filter_by(student_id=id).order_by(Attendance.date.desc()).all()
    total = len(records)
    present = len([r for r in records if r.status == 'Present'])
    absent = len([r for r in records if r.status == 'Absent'])
    late = len([r for r in records if r.status == 'Late'])
    percentage = round((present + late) / total * 100, 1) if total > 0 else 0
    return render_template('student_report.html', student=student, records=records,
                            total=total, present=present, absent=absent, late=late, percentage=percentage)


@app.route('/reports/student/<int:id>/export.csv')
@login_required
def export_student_report_csv(id):
    student = Student.query.get_or_404(id)
    records = Attendance.query.filter_by(student_id=id).order_by(Attendance.date.desc()).all()
    rows = [['Date', 'Subject Code', 'Subject Name', 'Status']]
    for r in records:
        rows.append([r.date.isoformat(), r.subject.code, r.subject.name, r.status])
    filename = f"attendance_{student.roll_no}_{date.today().isoformat()}.csv"
    return send_csv(rows, filename)


#--------------------------------------------------------------------------Users Management Routes
@app.route('/users')
@login_required
@admin_required
def users():
    all_users = User.query.all()
    return render_template('users.html', users=all_users)


@app.route('/users/delete/<int:id>')
@login_required
@admin_required
def delete_user(id):
    if id == session.get('user_id'):
        flash("You can't delete your own account.", 'danger')
        return redirect(url_for('users'))
    user = User.query.get_or_404(id)
    db.session.delete(user)
    db.session.commit()
    flash('User deleted.', 'success')
    return redirect(url_for('users'))

#-----------------------------------------------------------Frame Upload and Fingerprint Recognition Routes
@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify(ok=False, error="No file field"), 400
 
    file = request.files["file"]
 
    if file.filename == "":
        return jsonify(ok=False, error="Empty filename"), 400
 
    if not allowed_file(file.filename):
        return jsonify(ok=False, error="Unsupported file type"), 400
 
    filename = secure_filename(file.filename)
 
    if filename.lower().endswith(FIRST_FRAME_SUFFIX):
        clear_frame_directory()
 
    save_path = CAM_FRAME_PATH / filename
    file.save(str(save_path))
 
    frame_count = len(recognizer._get_frames())
    print(f"[UPLOAD] {frame_count}/{EXPECTED_FRAMES} frames")
 
    return jsonify(ok=True, status="saved", frames_received=frame_count), 200

    
@app.route("/fingerprint", methods=["POST"])
def fingerprint():
    data = request.get_json(silent=True) or {}
    fingerprint_roll_no = data.get("roll_no")
    subject_id = data.get("subject_id")   # required -- see note above about why
 
    if not fingerprint_roll_no:
        return jsonify(ok=False, error="missing 'roll_no'"), 400
 
    if not subject_id:
        return jsonify(ok=False, error="missing 'subject_id'"), 400
 
    print(f"[FINGERPRINT] roll_no={fingerprint_roll_no} subject_id={subject_id}")
 
    # Run recognition on whatever frames are currently in cam_frames/.
    result = recognizer.recognize()

    print("\nFACE RECOGNITION RESULT ")
    print("Result:", result)
    print("Status:", result.get("status"))
    print("Predicted Roll Number:", result.get("roll_number"))
    print("Votes:", result.get("votes"))
    print("Total Predictions:", result.get("total_predictions"))
    
    # result["status"] is "recognized" only when the face pipeline itself
    # found a confident majority match -- "waiting" (not enough frames yet)
    # and "invalid" both fall through to the mismatch/invalid path below.
    if result["status"] != "recognized":
        print(f"[FINGERPRINT] face pipeline did not recognize: {result}")
        clear_frame_directory()
        return jsonify(
            ok=True,
            result="invalid",
            reason="face_not_recognized",
            fingerprint_roll_number=fingerprint_roll_no,
            face_roll_number=None,
        ), 200
 
    face_roll_no = result["roll_number"]
 
    if face_roll_no != fingerprint_roll_no:
        print(f"[FINGERPRINT] mismatch: face={face_roll_no} fingerprint={fingerprint_roll_no}")
        clear_frame_directory()
        return jsonify(
            ok=True,
            result="invalid",
            reason="mismatch",
            fingerprint_roll_number=fingerprint_roll_no,
            face_roll_number=face_roll_no,
        ), 200
 
    # Match -- look up the student and mark attendance.
    student = Student.query.filter_by(roll_no=face_roll_no).first()
    if student is None:
        print(f"[FINGERPRINT] no Student record for roll_no={face_roll_no}")
        clear_frame_directory()
        return jsonify(
            ok=True,
            result="invalid",
            reason="roll_number_not_enrolled",
            fingerprint_roll_number=fingerprint_roll_no,
            face_roll_number=face_roll_no,
        ), 200
 
    today = date.today()
    existing = Attendance.query.filter_by(
        student_id=student.id, subject_id=subject_id, date=today
    ).first()
 
    if existing is None:
        db.session.add(Attendance(
            student_id=student.id,
            subject_id=subject_id,
            date=today,
            status="Present",
        ))
        db.session.commit()
        print(f"[FINGERPRINT] attendance marked present: {face_roll_no}")
    else:
        print(f"[FINGERPRINT] attendance already recorded today: {face_roll_no}")
 
    clear_frame_directory()
 
    return jsonify(
        ok=True,
        result="present",
        roll_number=face_roll_no,
        votes=result["votes"],
        total_predictions=result["total_predictions"],
    ), 200

#-----------------------------------------------------------------------------------Test Route
# @app.route("/test-recognition")
# def test_recognition():

#     result = recognizer.recognize()

#     print("Predicted Roll Number:", result.get("roll_number"))

#     return jsonify(result)

#--------------------------------------------------------------------------------INIT

def init_db():
    with app.app_context():
        db.create_all()
        if User.query.count() == 0:
            admin = User(username='admin', role='admin')
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print("Default admin created -> username: admin | password: admin123")


if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
